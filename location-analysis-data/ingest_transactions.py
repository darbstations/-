# -*- coding: utf-8 -*-
"""يبتلع ملفات معاملات درب الشهرية (xlsx) من مجلد (بأي تداخل) ويُخرج اللترات الفعلية يومياً لكل محطة.
Usage: python3 ingest_transactions.py <folder-or-zip> [more...] -> writes actual_daily.json"""
import sys, os, re, json, zipfile, tempfile
from collections import defaultdict
import openpyxl

CODE_RE = re.compile(r'([A-Z]{2}\d{3})')
FUEL_MAP = {'gasoline 91': 'l91', 'gasoline91': 'l91', 'gasoline 95': 'l95', 'gasoline95': 'l95', 'diesel': 'ld'}

def norm_date(v):
    """DateOfBusiness may be datetime, or text like 1/31/26."""
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip().split(' ')[0]
    m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{2,4})$', s)
    if m:
        mo, dy, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
        yr += 2000 if yr < 100 else 0
        return f'{yr:04d}-{mo:02d}-{dy:02d}'
    m = re.match(r'(\d{4})-(\d{2})-(\d{2})', s)
    return m.group(0) if m else None

def num(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    try: return float(str(v).replace(',', ''))
    except ValueError: return 0.0

def ingest_xlsx(path, agg, meta):
    m = CODE_RE.search(os.path.basename(path))
    if not m:
        meta['skipped_nocode'].append(os.path.basename(path)); return
    code = m.group(1)
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        meta['errors'].append(f'{os.path.basename(path)}: {e}'); return
    ws = None
    for name in wb.sheetnames:
        if 'transaction' in name.lower(): ws = wb[name]; break
    if ws is None: ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(h or '').strip() for h in next(rows)]
    except StopIteration:
        meta['errors'].append(f'{os.path.basename(path)}: empty'); return
    idx = {h.lower(): i for i, h in enumerate(header)}
    def col(*names):
        for n in names:
            if n.lower() in idx: return idx[n.lower()]
        return None
    c_date = col('DateOfBusiness', 'Date Of Business', 'Date')
    c_vol  = col('ResponseVolume', 'RequestVolume', 'Volume')
    c_amt  = col('ResponseAmount', 'RequestAmount', 'Amount')
    c_fuel = col('FuelName', 'Fuel')
    c_type = col('TransactionType')
    if c_date is None or c_vol is None:
        meta['errors'].append(f'{os.path.basename(path)}: missing columns {header[:8]}'); return
    n = 0
    for r in rows:
        if r is None or c_date >= len(r): continue
        d = norm_date(r[c_date])
        if not d: continue
        if c_type is not None and c_type < len(r) and r[c_type] and str(r[c_type]).strip().lower() not in ('sales', 'sale'):
            continue
        v = num(r[c_vol]) if c_vol < len(r) else 0.0
        a = num(r[c_amt]) if (c_amt is not None and c_amt < len(r)) else 0.0
        rec = agg[code][d]
        rec['lit'] += v; rec['rev'] += a; rec['n'] += 1
        if c_fuel is not None and c_fuel < len(r) and r[c_fuel]:
            k = FUEL_MAP.get(str(r[c_fuel]).strip().lower())
            if k: rec[k] += v
        n += 1
    meta['files'].append((os.path.basename(path), code, n))

def walk(target, agg, meta):
    if target.lower().endswith('.zip'):
        tmp = tempfile.mkdtemp(prefix='darbtx_')
        with zipfile.ZipFile(target) as z: z.extractall(tmp)
        target = tmp
    if os.path.isfile(target) and target.lower().endswith('.xlsx'):
        ingest_xlsx(target, agg, meta); return
    for root, _, files in os.walk(target):
        for f in files:
            if f.lower().endswith('.xlsx') and not f.startswith('~$'):
                ingest_xlsx(os.path.join(root, f), agg, meta)
            elif f.lower().endswith('.zip'):
                walk(os.path.join(root, f), agg, meta)

if __name__ == '__main__':
    agg = defaultdict(lambda: defaultdict(lambda: dict(lit=0.0, rev=0.0, n=0, l91=0.0, l95=0.0, ld=0.0)))
    meta = dict(files=[], errors=[], skipped_nocode=[])
    for t in sys.argv[1:]:
        walk(t, agg, meta)
    out = {c: {d: {k: round(v, 2) for k, v in rec.items()} for d, rec in days.items()} for c, days in agg.items()}
    json.dump(out, open('actual_daily.json', 'w'), ensure_ascii=False)
    print('stations:', len(out), '| files ok:', len(meta['files']), '| errors:', len(meta['errors']))
    for e in meta['errors'][:10]: print('ERR:', e)
    for c in sorted(out):
        days = out[c]
        print(c, '| days:', len(days), '| liters:', f"{sum(r['lit'] for r in days.values()):,.0f}",
              '| rev:', f"{sum(r['rev'] for r in days.values()):,.0f}", '| txns:', sum(r['n'] for r in days.values()))
