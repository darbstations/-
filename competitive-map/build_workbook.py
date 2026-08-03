# -*- coding: utf-8 -*-
"""يبني ملف الإكسل: بيانات المواقع + مصفوفة المسافات + جداول الحصر."""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

geo = json.load(open("geo.json", encoding="utf-8"))
S = geo["stations"]
pois = json.load(open("pois.json", encoding="utf-8")) if os.path.exists("pois.json") else {}

AMBER = "E8760C"; DARK = "2E1D12"; PAPER = "F3E9DC"; GOLD = "C79A3A"
HDR  = Font(name="Tahoma", size=10, bold=True, color="FFFFFF")
BODY = Font(name="Tahoma", size=10)
TITLE= Font(name="Tahoma", size=13, bold=True, color=DARK)
FILL = PatternFill("solid", fgColor=AMBER)
SUB  = PatternFill("solid", fgColor=PAPER)
THIN = Side(style="thin", color="D9C6AE")
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CEN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT= Alignment(horizontal="right", vertical="center")

wb = Workbook()

def sheet(name, widths, headers, rows, title=None, note=None):
    ws = wb.create_sheet(name)
    ws.sheet_view.rightToLeft = True
    r = 1
    if title:
        ws.cell(1, 1, title).font = TITLE
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.cell(1, 1).alignment = RIGHT
        ws.row_dimensions[1].height = 24
        r = 3
    hr = r
    for c, h in enumerate(headers, 1):
        cell = ws.cell(hr, c, h)
        cell.font, cell.fill, cell.alignment, cell.border = HDR, FILL, CEN, BOX
    ws.row_dimensions[hr].height = 30
    for i, row in enumerate(rows, hr + 1):
        for c, v in enumerate(row, 1):
            cell = ws.cell(i, c, v)
            cell.font, cell.border = BODY, BOX
            cell.alignment = CEN if not isinstance(v, str) or len(str(v)) < 26 else RIGHT
            if i % 2 == 0: cell.fill = SUB
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = ws.cell(hr + 1, 1)
    if note:
        nr = hr + len(rows) + 2
        ws.cell(nr, 1, note).font = Font(name="Tahoma", size=9, italic=True, color="6E5A48")
        ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=len(headers))
        ws.cell(nr, 1).alignment = RIGHT
    return ws

# ---------- 1) المحطات ----------
rows = []
for s in S:
    lm = {l["name"]: l for l in s["landmarks"]}
    rows.append([
        s["code"], s["name"], geo["city"], s["lat"], s["lng"],
        s["nearest"], s["nearest_km"], s["overlap"]["1"], s["overlap"]["3"], s["overlap"]["5"],
        lm["المسجد الحرام"]["km"], lm["المسجد الحرام"]["dir"],
        lm["مشعر منى"]["km"], lm["مشعر مزدلفة"]["km"], lm["صعيد عرفات"]["km"],
        "ضمن تجمّع متنافس" if s["nearest_km"] < 5 else "موقع منفرد النطاق",
        s["maps_url"],
    ])
sheet("المحطات",
      [10, 18, 12, 13, 13, 10, 12, 11, 11, 11, 12, 12, 11, 12, 12, 18, 42],
      ["الرمز","اسم المحطة","المدينة","خط العرض","خط الطول","أقرب محطة","المسافة (كم)",
       "تداخل ١كم %","تداخل ٣كم %","تداخل ٥كم %","الحرم (كم)","اتجاه الحرم",
       "منى (كم)","مزدلفة (كم)","عرفات (كم)","التصنيف التنافسي","رابط الخريطة"],
      rows, title="محطات درب — مكة المكرمة · بيانات الموقع والتحليل التنافسي",
      note="نسبة التداخل = نسبة تقاطع دائرة نطاق الخدمة مع دائرة أقرب محطة، بنفس نصف القطر.")

# ---------- 2) مصفوفة المسافات ----------
codes = [s["code"] for s in S]
mrows = []
for a in S:
    row = [f'{a["name"]} ({a["code"]})']
    for c in codes:
        row.append("—" if c == a["code"] else geo["matrix"][a["code"] + "|" + c])
    mrows.append(row)
sheet("مصفوفة المسافات", [26] + [15] * len(codes),
      ["من \\ إلى"] + [f'{s["name"]}\n{s["code"]}' for s in S], mrows,
      title="المسافات البينية بين المواقع (كم — مسافة مباشرة)",
      note="القيم أقل من ٥ كم تعني تداخلاً فعلياً في نطاق الخدمة بين المحطتين.")

# ---------- 3-5) جداول الحصر ----------
INV = [
    ("تأجير السيارات", "حصر مواقع تأجير السيارات المحيطة بكل محطة",
     ["المحطة","رمز المحطة","اسم المنشأة","النشاط","العنوان","خط العرض","خط الطول",
      "المسافة عن المحطة (كم)","الاتجاه","الهاتف","الموقع الإلكتروني","التقييم","عدد المراجعات","ملاحظات"]),
    ("مكاتب الحج والعمرة", "حصر مكاتب وحملات الحج والعمرة المحيطة بكل محطة",
     ["المحطة","رمز المحطة","اسم المكتب/الحملة","النوع","العنوان","خط العرض","خط الطول",
      "المسافة عن المحطة (كم)","الاتجاه","الهاتف","الموقع الإلكتروني","التقييم","عدد المراجعات","ملاحظات"]),
    ("الشركات الحكومية والخاصة", "حصر الجهات الحكومية وشركات القطاع الخاص المحيطة بكل محطة",
     ["المحطة","رمز المحطة","اسم الجهة","القطاع (حكومي/خاص)","النشاط","العنوان","خط العرض","خط الطول",
      "المسافة عن المحطة (كم)","الاتجاه","الهاتف","الموقع الإلكتروني","حجم الأسطول التقديري","ملاحظات"]),
]
W = [20, 11, 30, 18, 34, 12, 12, 16, 11, 16, 28, 9, 12, 26]
for name, title, headers in INV:
    data = []
    for s in S:
        for p in (pois.get(s["code"], {}).get(name, []) or []):
            data.append([s["name"], s["code"]] + p)
    note = ("" if data else
            "⚠ لم تُجمَع البيانات بعد: حساب Apify تجاوز الحد الشهري، ونطاق overpass-api.de محجوب بسياسة الشبكة. "
            "شغّل collect_pois.py ثم build_workbook.py لتعبئة هذا الجدول تلقائياً.")
    sheet(name, W[:len(headers)], headers, data, title=title, note=note)

# ---------- 6) المنهجية وحالة البيانات ----------
ws = wb.create_sheet("المنهجية وحالة البيانات")
ws.sheet_view.rightToLeft = True
ws.column_dimensions["A"].width = 30; ws.column_dimensions["B"].width = 95
lines = [
    ("المصدر الأساسي", "ملف الإحداثيات المرفق — ٥ محطات بمكة المكرمة (الرمز، الاسم، الإحداثيات)."),
    ("حساب المسافات", "معادلة Haversine على نصف قطر أرضي 6371.0088 كم — مسافة مباشرة وليست مسافة طريق."),
    ("نسبة التداخل", "مساحة تقاطع دائرتين متساويتي نصف القطر ÷ مساحة الدائرة الواحدة × ١٠٠."),
    ("المعالم المرجعية", "المسجد الحرام ومشاعر منى ومزدلفة وعرفات — إحداثيات مرجعية ثابتة للتوجيه."),
    ("حالة الحصر الميداني", "غير مكتمل. تعذّر سحب بيانات المنشآت المحيطة لسببين:"),
    ("  السبب الأول", "حساب Apify: تجاوز الحد الشهري للاستخدام (Monthly usage hard limit exceeded)."),
    ("  السبب الثاني", "نطاق overpass-api.de (OpenStreetMap) محجوب بسياسة الشبكة للجلسة (403)."),
    ("خطوة الإكمال", "ارفع حد استخدام Apify ثم شغّل: python3 collect_pois.py && python3 build_workbook.py && python3 build_report.py"),
    ("نطاق البحث المقترح", "٥ كم حول كل محطة، بمصطلحات: تأجير سيارات · حج وعمرة · شركة · حكومي · محطة وقود."),
    ("ما لم يُدرَج", "لم تُدرَج أي أسماء منشآت أو إحداثيات تقديرية — الجداول فارغة عمداً بدلاً من تعبئتها ببيانات غير موثقة."),
]
ws.cell(1, 1, "منهجية الدراسة وحالة البيانات").font = TITLE
ws.merge_cells("A1:B1")
for i, (k, v) in enumerate(lines, 3):
    a = ws.cell(i, 1, k); b = ws.cell(i, 2, v)
    a.font = Font(name="Tahoma", size=10, bold=True, color=DARK); a.fill = SUB
    b.font = BODY
    for c in (a, b): c.alignment = RIGHT; c.border = BOX
    ws.row_dimensions[i].height = 22

wb.remove(wb["Sheet"])
wb.save("darb-competitive-map-makkah.xlsx")
print("workbook saved · sheets:", wb.sheetnames)
