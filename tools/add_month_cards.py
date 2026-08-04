# -*- coding: utf-8 -*-
"""بطاقة تحليل لكل شهر بأسلوب البيرسونا: أسباب الصعود والهبوط مستخرجة من الأرقام."""
import openpyxl, re, html, collections

SRC = OUT = "/home/user/-/darb-five-stations-analysis.html"
BUDGET = "/root/.claude/uploads/447348d0-0f0b-5d32-9f6b-27c9ad645473/994d4b5d-________Sales_Analysis_2026_Actual_vs_Budget_30.04.2026.xlsx"
CSFILE = "/root/.claude/uploads/447348d0-0f0b-5d32-9f6b-27c9ad645473/a16169b5-Monthly_Report_2026.xlsx"
KEEP = ["MK007", "MK017", "MK002", "MK023", "MK019"]
MON = ["يناير", "فبراير", "مارس", "أبريل", "مايو", "يونيو"]
CS_SHEETS = ["Jan", "Feb", "Mar", "Apr", "May", "June"]

E = lambda t: html.escape(str(t), quote=True)
F = lambda n: f"{round(n):,}"
def AB(n):
    n = round(n)
    return f"{n/1_000_000:.1f}م" if n >= 1_000_000 else (f"{round(n/1000):,}ألف" if n >= 1000 else f"{n:,}")
def P(x):                       # نسبة بإشارة
    return f"{x:+.0f}٪"

#  أسباب مرشّحة — طبقة فرضيات لا قياس، تُعلَّم بالضغط وتُحرَّر
HYPO = ["موسم العمرة والحج", "رمضان والعيد", "إجازة مدارس", "إغلاق أو تحويلة طريق",
        "صيانة مضخات أو توقف جزئي", "انقطاع منتج", "منافس جديد قريب", "تغيّر أسعار",
        "حملة تسويقية", "تغيّر فريق التشغيل", "طقس أو أمطار", "أعمال إنشائية مجاورة"]

# ═══════════ 1. الموازنة و2025 ═══════════
wb = openpyxl.load_workbook(BUDGET, data_only=True)
rows = list(wb["Sales Analysis 2026"].iter_rows(values_only=True))
N0 = lambda v: v if isinstance(v, (int, float)) else 0
BU = {}
for i in range(2, 970, 11):
    b = rows[i:i + 11]
    if not b[0][1]:
        continue
    d = {}
    for r in b:
        d[str(r[5]).strip()] = [N0(r[6 + j]) for j in range(12)]
    BU[str(b[0][2]).strip()] = d

# ═══════════ 2. خدمة العملاء شهريًا ═══════════
wb2 = openpyxl.load_workbook(CSFILE, read_only=True, data_only=True)
def load(n):
    r = list(wb2[n].iter_rows(values_only=True))
    h = [str(c).strip() if c else "" for c in r[0]]
    o = []
    for x in r[1:]:
        if not any(c is not None and str(c).strip() not in ("", "-") for c in x):
            continue
        d = {}
        for i, k in enumerate(h):
            if i < len(x) and k and k not in d:
                d[k] = x[i]
        o.append(d)
    return o
CS = {c: [dict(cnt=0, comp=0, top=None) for _ in range(6)] for c in KEEP}
for j, sh in enumerate(CS_SHEETS):
    cats = {c: collections.Counter() for c in KEEP}
    for d in load(sh):
        m = re.search(r"\b([A-Z]{2}\d{2,4})\b", str(d.get("station") or "").upper().replace(" ", ""))
        c = m.group(1) if m else ""
        if c not in CS:
            continue
        CS[c][j]["cnt"] += 1
        if str(d.get("Reporting Entity") or "").strip().lower() == "complaint":
            CS[c][j]["comp"] += 1
        sub = str(d.get("Subcategory") or "").strip()
        if sub and sub != "-":
            cats[c][sub] += 1
    for c in KEEP:
        if cats[c]:
            CS[c][j]["top"] = cats[c].most_common(1)[0]

# ═══════════ 3. الجدول الشهري من التقرير ═══════════
src = open(SRC, encoding="utf-8").read()
pages_head, pages_all = src.split('<main class="wrap" id="pages">', 1)
blocks = re.split(r'(?=<div class="pgview")', pages_all)
lead, blocks = blocks[0], blocks[1:]
PG, ORDER = {}, []
for b in blocks:
    k = re.search(r'id="pg-([\w-]+)"', b).group(1)
    PG[k] = b
    ORDER.append(k)

Nn = lambda x: int(str(x).replace(",", ""))
MONTHS = {}
for c in KEEP:
    tb = re.search(r'<th>الشهر</th>.*?<tbody>(.*?)</tbody>', PG[c + "-monthly"], re.S).group(1)
    rr = re.findall(r'<tr><td><b>([^<]+)</b></td><td>(\d+)</td><td>([\d,]+)</td>\s*<td>([\d,]+)</td>'
                    r'<td>([\d,]+)</td><td>([\d,]+)</td>\s*<td>([\d,]+)</td><td>([\d,]+)</td><td>([^<]*)</td>', tb)
    lst = []
    for j, r in enumerate(rr):
        lst.append(dict(m=r[0], days=int(r[1]), rev=Nn(r[2]), vis=Nn(r[3]), lit=Nn(r[4]),
                        inv=Nn(r[5]), revd=Nn(r[6]), litd=Nn(r[7]), peak=r[8].strip(),
                        visd=Nn(r[3]) / int(r[1]),
                        bu=BU[c]["BU.2026"][j], act=BU[c]["Act.2026"][j], a25=BU[c]["Act.2025"][j],
                        cs=CS[c][j]))
    MONTHS[c] = lst

# ═══════════ 4. توليد الأسباب من الأرقام ═══════════
def build_reasons(cur, prev, all_):
    up, dn = [], []
    revs = [x["revd"] for x in all_]
    if cur["revd"] == max(revs):
        up.append("<b>أعلى شهر</b> في النصف الأول بمتوسط الإيراد اليومي")
    if cur["revd"] == min(revs):
        dn.append("<b>أدنى شهر</b> في النصف الأول بمتوسط الإيراد اليومي")

    if prev:
        dr = cur["revd"] / prev["revd"] - 1
        dv = cur["visd"] / prev["visd"] - 1
        di = cur["inv"] / prev["inv"] - 1
        #  تفكيك التغير: كم منه من عدد الزيارات وكم من قيمة الفاتورة
        tot = cur["revd"] - prev["revd"]
        vis_eff = (cur["visd"] - prev["visd"]) * prev["inv"]
        share = abs(vis_eff) / (abs(vis_eff) + abs(cur["visd"] * (cur["inv"] - prev["inv"]))) * 100 \
            if (cur["visd"] - prev["visd"]) or (cur["inv"] - prev["inv"]) else 0
        line = (f'الإيراد اليومي {P(dr*100)} عن {prev["m"]} '
                f'({F(prev["revd"])} ← {F(cur["revd"])} ر.س)')
        (up if dr >= 0 else dn).append(line)
        if abs(dv) >= 0.02:
            (up if dv > 0 else dn).append(
                f'الزيارات اليومية {P(dv*100)} ({F(prev["visd"])} ← {F(cur["visd"])} سيارة/يوم)')
        if abs(di) >= 0.02:
            (up if di > 0 else dn).append(
                f'متوسط الفاتورة {P(di*100)} ({prev["inv"]} ← {cur["inv"]} ر.س) — '
                + ("سلة أكبر لكل عميل" if di > 0 else "تعبئة أصغر لكل عميل"))
        if tot and share:
            drv = "عدد الزيارات" if share >= 50 else "قيمة الفاتورة"
            (up if tot > 0 else dn).append(
                f'<b>مصدر التغير:</b> نحو {share if share>=50 else 100-share:.0f}٪ منه يعود إلى '
                f'<b>{drv}</b> لا إلى الطرف الآخر')
        if cur["peak"] != prev["peak"]:
            (up if dr >= 0 else dn).append(
                f'ساعة الذروة تحوّلت من {prev["peak"]} إلى {cur["peak"]} — تغيّر في نوع الحركة')

    if cur["bu"]:
        ach = cur["act"] / cur["bu"] * 100
        txt = f'الإنجاز مقابل الموازنة <b>{ach:.0f}٪</b> ({AB(cur["act"])} لتر مقابل {AB(cur["bu"])})'
        if prev and prev["bu"]:
            txt += f' — بعد {prev["act"]/prev["bu"]*100:.0f}٪ في {prev["m"]}'
        (up if ach >= 100 else dn).append(txt)
    if cur["a25"]:
        g = cur["act"] / cur["a25"] - 1
        (up if g >= 0 else dn).append(
            f'{P(g*100)} مقابل {cur["m"]} 2025 ({AB(cur["a25"])} ← {AB(cur["act"])} لتر)')

    cs, pcs = cur["cs"], prev["cs"] if prev else None
    if cs["cnt"]:
        t = f'{cs["cnt"]} سجل خدمة عملاء ({cs["comp"]} شكوى)'
        if cs["top"]:
            t += f' — أبرزها «{cs["top"][0]}» ({cs["top"][1]})'
        if pcs and pcs["cnt"] and cs["cnt"] > pcs["cnt"]:
            dn.append(t + f' · ارتفاع من {pcs["cnt"]} في {prev["m"]}')
        elif pcs and cs["cnt"] < pcs["cnt"]:
            up.append(t + f' · انخفاض من {pcs["cnt"]} في {prev["m"]}')
        else:
            dn.append(t)
    elif prev and prev["cs"]["cnt"]:
        up.append(f'لا سجلات خدمة عملاء هذا الشهر — بعد {prev["cs"]["cnt"]} في {prev["m"]}')

    if cur["days"] < 30:
        dn.append(f'{cur["days"]} يومًا مسجلًا فقط — أثر ميكانيكي على الإجمالي الشهري '
                  '(كل المقارنات أعلاه على متوسط اليوم لتحييده)')
    return up, dn

def state_of(cur):
    if cur["bu"]:
        a = cur["act"] / cur["bu"] * 100
        if a >= 110: return "ممتاز", "c-nbh"
        if a >= 95:  return "جيد", "c-hwy"
        if a >= 80:  return "يحتاج متابعة", "c-rem"
        return "حرج", "c-viv"
    return "غير مصنّف", "c-un"

# ═══════════ 5. بناء البطاقات ═══════════
ez = max(int(x) for x in re.findall(r'data-ez="z(\d+)"', src))

for c in KEEP:
    ms = MONTHS[c]
    cards = []
    for j, cur in enumerate(ms):
        prev = ms[j - 1] if j else None
        up, dn = build_reasons(cur, prev, ms)
        st, stc = state_of(cur)
        dr = (cur["revd"] / prev["revd"] - 1) * 100 if prev else 0
        ico = "📈" if dr > 2 else ("📉" if dr < -2 else "➖")
        ach = f'{cur["act"]/cur["bu"]*100:.0f}٪ إنجاز' if cur["bu"] else "بلا موازنة"
        head = (f'{ico} <b>{cur["m"]}</b>'
                f'<span class="msum">{ach} · {P(dr) if prev else "أول شهر"} عن الشهر السابق '
                f'· {F(cur["revd"])} ر.س/يوم</span>'
                f'<span class="cls {stc} mstate" title="اضغط لتغيير الحالة">{st}</span>')
        chips = "".join(f'<span class="hchip">{E(x)}</span>' for x in HYPO)
        cards.append(
            f'<details class="mdet"{" open" if j == len(ms)-1 else ""}>'
            f'<summary>{head}</summary>'
            '<div class="mbody">'
            '<div class="mnums">'
            f'<span>إيراد/يوم <b>{F(cur["revd"])}</b> ر.س</span>'
            f'<span>زيارات/يوم <b>{F(cur["visd"])}</b></span>'
            f'<span>فاتورة <b>{cur["inv"]}</b> ر.س</span>'
            f'<span>لترات <b>{AB(cur["lit"])}</b></span>'
            f'<span>لترات/يوم <b>{F(cur["litd"])}</b></span>'
            f'<span>ذروة <b>{cur["peak"]}</b></span>'
            f'<span>أيام مسجلة <b>{cur["days"]}</b></span>'
            "</div>"
            + (f'<div class="mline up"><b>▲ ما دفع للأعلى</b><ul>'
               + "".join(f"<li>{x}</li>" for x in up) + "</ul></div>" if up else "")
            + (f'<div class="mline dn"><b>▼ ما ضغط للأسفل</b><ul>'
               + "".join(f"<li>{x}</li>" for x in dn) + "</ul></div>" if dn else "")
            + '<div class="mline hyp"><b>أسباب مرشّحة للتحقق</b>'
              '<div class="mnote">لم تُقَس — اضغط على ما ينطبق ليُعلَّم، وأضف غيره من وضع التحرير</div>'
              f'<div class="hchips">{chips}</div></div>'
            '<div class="mline act"><b>السبب المؤكد والإجراء</b>'
            '<p class="mfree">اكتب هنا ما تعرفه عن سبب حركة هذا الشهر، والإجراء المتخذ…</p></div>'
            "</div></details>")

    ez += 1
    section = (
        f'<div class="sec-h" style="margin-top:20px" data-ez="z{ez}"><h2>تحليل كل شهر</h2>'
        '<span>أسباب الصعود والهبوط مستخرجة من أرقام الشهر — اضغط على الشهر لفتحه</span></div>'
        f'<div class="mcards" data-ez="z{ez+1}">{"".join(cards)}</div>'
        '<div class="dnote">📐 نقاط <b>▲</b> و<b>▼</b> مشتقة آليًا من مقارنة أرقام الشهر بالشهر السابق '
        'وبالموازنة وبنفس الشهر من 2025 وبسجلات خدمة العملاء — كلها <b>مقيسة</b>. '
        'شرائح «أسباب مرشّحة» و«السبب المؤكد» <b>غير مقيسة</b> ومتروكة لكم؛ الأرقام تقول '
        '<i>ما الذي</i> تغيّر ومن أي طرف (زيارات أم فاتورة)، ولا تقول <i>لماذا</i>.</div>')
    ez += 1
    p = PG[c + "-monthly"]
    anchor = '<div class="sec-h" style="margin-top:18px"><h2>مزيج الوقود شهريًا</h2>'
    assert anchor in p, c
    PG[c + "-monthly"] = p.replace(anchor, section + anchor, 1)

pages_all = lead + "".join(PG[k] for k in ORDER)

# ═══════════ 6. الأنماط والسلوك ═══════════
CSS = """
<style id="mcards-css">
/* ── بطاقات تحليل الشهر ── */
.mcards{display:grid;gap:10px;margin-bottom:16px}
.mdet{background:var(--card);border:1px solid var(--line);border-radius:14px;
  box-shadow:var(--shadow);overflow:hidden}
.mdet>summary{cursor:pointer;padding:13px 16px;display:flex;align-items:center;gap:10px;
  flex-wrap:wrap;font-size:15px;list-style:none}
.mdet>summary::-webkit-details-marker{display:none}
.mdet>summary::after{content:"▾";margin-inline-start:auto;color:var(--ink3);font-size:13px}
.mdet[open]>summary{border-bottom:1px dashed var(--line2);background:#FBF9F5}
.mdet[open]>summary::after{content:"▴"}
.mdet>summary b{font-size:16px;font-weight:800}
.mdet .msum{font-size:12.5px;color:var(--ink2);font-family:'Tajawal'}
.mdet .mstate{cursor:pointer;user-select:none;-webkit-user-select:none}
.mbody{padding:14px 16px 15px}
.mnums{display:flex;flex-wrap:wrap;gap:8px;margin-bottom:13px}
.mnums span{background:#FDFCFA;border:1px solid var(--line);border-radius:10px;
  padding:7px 11px;font-size:12px;color:var(--ink2)}
.mnums b{font-size:14px;font-weight:800;color:var(--ink)}
.mline{border-radius:11px;padding:10px 13px;margin-bottom:9px;font-size:12.5px;line-height:1.9}
.mline b{font-size:13px}
.mline ul{padding-inline-start:17px;margin-top:5px;display:grid;gap:4px;color:var(--ink2)}
.mline.up{background:#F0F7F2;border:1px solid #CBE2D3}.mline.up>b{color:#22694F}
.mline.dn{background:#FBF0ED;border:1px solid #EBCFC7}.mline.dn>b{color:#A6432E}
.mline.hyp{background:#F8F6F2;border:1px dashed var(--line2)}
.mline.act{background:#FBF6EF;border:1px solid #EAD9C3}.mline.act>b{color:#8A5A2B}
.mnote{font-size:11px;color:var(--ink3);margin:2px 0 7px}
.hchips{display:flex;flex-wrap:wrap;gap:6px}
.hchip{font-size:11.5px;background:#fff;border:1px solid var(--line2);border-radius:8px;
  padding:3px 9px;color:var(--ink3);cursor:pointer;transition:.14s;user-select:none;-webkit-user-select:none}
.hchip:hover{border-color:var(--orange);color:var(--ink2)}
.hchip.on{background:var(--orange);border-color:var(--orange);color:#fff;font-weight:700}
.hchip.on::before{content:"✓ "}
.mfree{color:var(--ink3);font-style:italic;margin-top:4px}
.mfree.filled{color:var(--ink);font-style:normal}
@media(max-width:620px){.mdet>summary{font-size:13.5px}.mnums span{font-size:11.5px}}
</style>
"""

JS = r"""
<script id="mcards-js">
/* ═══ سلوك بطاقات الشهر: تعليم الأسباب المرشّحة وتدوير وسم الحالة ═══ */
(function(){
  var STATES=[['ممتاز','c-nbh'],['جيد','c-hwy'],['يحتاج متابعة','c-rem'],['حرج','c-viv'],['غير مصنّف','c-un']];
  function fire(el){var z=el.closest('[data-ez]');if(z)z.dispatchEvent(new Event('input',{bubbles:true}));}

  document.addEventListener('click',function(e){
    var chip=e.target.closest&&e.target.closest('.hchip');
    if(chip){chip.classList.toggle('on');fire(chip);return;}
    var st=e.target.closest&&e.target.closest('.mstate');
    if(st){
      e.preventDefault();
      var i=STATES.findIndex(function(s){return st.classList.contains(s[1]);});
      var n=STATES[(i+1)%STATES.length];
      STATES.forEach(function(s){st.classList.remove(s[1]);});
      st.classList.add(n[1]); st.textContent=n[0]; fire(st);
    }
  });

  /* النص الإرشادي يختفي أول ما يُكتب فوقه */
  document.addEventListener('focusin',function(e){
    var f=e.target.closest&&e.target.closest('.mfree');
    if(f&&!f.classList.contains('filled')){f.textContent='';f.classList.add('filled');}
  });
  document.addEventListener('focusout',function(e){
    var f=e.target.closest&&e.target.closest('.mfree');
    if(f&&!f.textContent.trim()){
      f.classList.remove('filled');
      f.textContent='اكتب هنا ما تعرفه عن سبب حركة هذا الشهر، والإجراء المتخذ…';
    }
  });
})();
</script>
"""

doc = pages_head + '<main class="wrap" id="pages">' + pages_all
doc = doc.replace("</head>", CSS + "</head>", 1)
doc = doc.replace("</body>", JS + "</body>", 1)
doc = re.sub(r'data-docid="[^"]*"', 'data-docid="darb-5st-months-v1"', doc, count=1)
open(OUT, "w", encoding="utf-8").write(doc)
print("تم · الحجم:", round(len(doc.encode()) / 1024), "KB · مناطق تحرير حتى z%d" % ez)
