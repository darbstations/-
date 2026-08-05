# -*- coding: utf-8 -*-
"""تشغيل استخراج الشركاء المحتملين حول المحطات الخمس عبر Apify، ثم تجميع النتائج.

يعمل بمجرد توفّر رصيد في حساب Apify. يشغّل أكتور compass/crawler-google-places
مرة لكل محطة (الأكتور يقبل نطاقًا جغرافيًا واحدًا لكل تشغيل)، ثم يوحّد المخرجات
ويصنّفها ويحسب المسافة من المحطة.

    export APIFY_TOKEN=...
    python3 tools/apify_nearby_partners.py            # يشغّل ويجمّع
    python3 tools/apify_nearby_partners.py --collect  DATASET_ID...   # تجميع فقط
"""
import json, math, os, sys, time, urllib.request, urllib.parse

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SPEC = os.path.join(BASE, "apify", "all-runs.json")
OUT_JSON = os.path.join(BASE, "apify", "nearby-partners.json")
OUT_XLSX = os.path.join(BASE, "darb-nearby-partners.xlsx")
TOKEN = os.environ.get("APIFY_TOKEN", "")
API = "https://api.apify.com/v2"

spec = json.load(open(SPEC, encoding="utf-8"))
ACTOR = spec["actor"].replace("/", "~")
CATS = spec["categories"]
TERM2CAT = {t: c for c, ts in CATS.items() for t in ts}


def api(path, method="GET", body=None):
    url = f"{API}{path}{'&' if '?' in path else '?'}token={TOKEN}"
    data = json.dumps(body).encode() if body is not None else None
    req = urllib.request.Request(url, data=data, method=method,
                                 headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=120) as r:
        return json.loads(r.read().decode())


def haversine(a_lat, a_lng, b_lat, b_lng):
    R = 6371000
    p1, p2 = math.radians(a_lat), math.radians(b_lat)
    dp = math.radians(b_lat - a_lat)
    dl = math.radians(b_lng - a_lng)
    h = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return round(2 * R * math.asin(math.sqrt(h)))


def run_all():
    """يبدأ تشغيلًا لكل محطة ويعيد قائمة (كود المحطة، معرّف الداتاست)."""
    started = []
    for r in spec["runs"]:
        res = api(f"/acts/{ACTOR}/runs", "POST", r["input"])["data"]
        print(f"  بدأ {r['code']}: run={res['id']}")
        started.append((r, res["id"]))
    out = []
    for r, run_id in started:
        while True:
            d = api(f"/actor-runs/{run_id}")["data"]
            if d["status"] in ("SUCCEEDED", "FAILED", "ABORTED", "TIMED-OUT"):
                break
            time.sleep(15)
        print(f"  {r['code']}: {d['status']}")
        if d["status"] == "SUCCEEDED":
            out.append((r, d["defaultDatasetId"]))
    return out


def fetch(ds):
    items, off = [], 0
    while True:
        q = urllib.parse.urlencode({"limit": 1000, "offset": off, "clean": "true"})
        chunk = api(f"/datasets/{ds}/items?{q}")
        if not chunk:
            break
        items += chunk
        off += len(chunk)
        if len(chunk) < 1000:
            break
    return items


def collect(pairs):
    rows, seen = [], set()
    for r, ds in pairs:
        for p in fetch(ds):
            loc = p.get("location") or {}
            if not loc.get("lat"):
                continue
            key = (r["code"], p.get("placeId"))
            if key in seen:
                continue
            seen.add(key)
            rows.append({
                "كود المحطة": r["code"], "المحطة": r["name"],
                "التصنيف لدينا": TERM2CAT.get(p.get("searchString", ""), "أخرى"),
                "مصطلح البحث": p.get("searchString"),
                "الاسم": p.get("title"), "فئة جوجل": p.get("categoryName"),
                "المسافة (م)": haversine(r["lat"], r["lng"], loc["lat"], loc["lng"]),
                "التقييم": p.get("totalScore"), "المراجعات": p.get("reviewsCount"),
                "الهاتف": p.get("phoneUnformatted") or p.get("phone"),
                "الموقع الإلكتروني": p.get("website"),
                "العنوان": p.get("address"), "الحي": p.get("neighborhood"),
                "رابط الخريطة": p.get("url"),
                "lat": loc["lat"], "lng": loc["lng"],
            })
    rows.sort(key=lambda x: (x["كود المحطة"], x["التصنيف لدينا"], x["المسافة (م)"]))
    json.dump(rows, open(OUT_JSON, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    write_xlsx(rows)
    print(f"\n{len(rows)} منشأة · {OUT_JSON} · {OUT_XLSX}")
    for c in dict.fromkeys(x["كود المحطة"] for x in rows):
        sub = [x for x in rows if x["كود المحطة"] == c]
        by = {}
        for x in sub:
            by[x["التصنيف لدينا"]] = by.get(x["التصنيف لدينا"], 0) + 1
        print(f"  {c}: {len(sub)} — " + " · ".join(f"{k} {v}" for k, v in by.items()))


def write_xlsx(rows):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    cols = [k for k in rows[0] if k not in ("lat", "lng")] if rows else []
    groups = [("الكل", rows)] + [(c, [x for x in rows if x["كود المحطة"] == c])
                                 for c in dict.fromkeys(x["كود المحطة"] for x in rows)]
    for title, data in groups:
        ws = wb.create_sheet(title[:31]); ws.sheet_view.rightToLeft = True
        ws.append(cols)
        for j in range(1, len(cols) + 1):
            c = ws.cell(1, j)
            c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="55565A")
            c.alignment = Alignment(horizontal="center")
        for x in data:
            ws.append([x[k] for k in cols])
        for j, k in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(j)].width = 34 if k in ("الاسم", "العنوان", "الموقع الإلكتروني", "رابط الخريطة") else 15
        ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    wb.save(OUT_XLSX)


if __name__ == "__main__":
    if not TOKEN:
        sys.exit("ضع رمز الوصول في APIFY_TOKEN")
    if len(sys.argv) > 2 and sys.argv[1] == "--collect":
        ids = sys.argv[2:]
        collect(list(zip(spec["runs"], ids)))
    else:
        print(f"تشغيل {len(spec['runs'])} مهمة على {spec['actor']} (نطاق {spec['radiusKm']} كم)…")
        collect(run_all())
