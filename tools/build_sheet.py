# -*- coding: utf-8 -*-
"""The campaign register as a workbook Google Sheets opens as its own file.

What the page does with script, the sheets do with formulas: the summary reads
the campaigns sheet, and the plan subtracts each budget from the partner's
total and shows what is left after every campaign. Dropdowns carry the same
lists the page offers, and the station table travels with it so codes stay
pickable.
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

D = json.load(open('sheet_src.json', encoding='utf-8'))
CAMPS = sorted(D['campaigns'], key=lambda c: (c.get('start') or '9999', c['code']))
STA = {s['code']: s for s in D['stations']}
TOOLS = D['tools']
PTYPES = [p for p in D['ptypes'] if not p.startswith('—')]
STATUSES = ['بانتظار الرد', 'موافق', 'معتذر', 'قيد التفاوض']

DEEP = '3A2416'
ORANGE = 'C85E05'
SAND = 'F1E2CC'
LINE = 'E0D0BA'
HOT = 'FBF0E9'

HEAD = Font(name='Arial', size=11, bold=True, color='FFFFFF')
BODY = Font(name='Arial', size=11)
BOLD = Font(name='Arial', size=11, bold=True)
ORNG = Font(name='Arial', size=11, bold=True, color=ORANGE)
thin = Side(style='thin', color=LINE)
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(horizontal='right', vertical='top', wrap_text=True)
MID = Alignment(horizontal='right', vertical='center')

wb = Workbook()


def sheet(title, headers, widths):
    ws = wb.create_sheet(title)
    ws.sheet_view.rightToLeft = True
    ws.append(headers)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in ws[1]:
        c.font = HEAD
        c.fill = PatternFill('solid', fgColor=DEEP)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BOX
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 30
    return ws


def dress(ws, first=2, max_col=None):
    for row in ws.iter_rows(min_row=first, max_col=max_col or ws.max_column):
        for c in row:
            c.font = BODY
            c.alignment = WRAP
            c.border = BOX


def stname(code):
    s = STA.get(code)
    return s['name'] if s else ''


def stcity(code):
    s = STA.get(code)
    return s.get('city', '') if s else ''


def dur(c):
    a, b = c.get('start') or '', c.get('end') or ''
    if c.get('days'):
        return '%s أيام' % c['days']
    if c.get('weeks'):
        return '%s أسابيع' % c['weeks']
    return ('%s ← %s' % (a, b)) if a and b else ''


# ---------------------------------------------------------------- الحملات
CH = ['كود الحملة', 'اسم الحملة', 'الحملة الأساسية', 'تاريخ البدء', 'تاريخ الانتهاء',
      'المدة', 'المحطات', 'عدد المحطات', 'الأدوات المستخدمة', 'الهدف', 'مؤشر القياس',
      'المستهدف', 'الميزانية (ر.س)', 'الشروط', 'العرض المقترح', 'عدد الشركاء',
      'ملاحظات', 'آخر تعديل']
ws = sheet('الحملات', CH, [15, 34, 15, 12, 12, 12, 34, 10, 34, 40, 26, 22, 14, 40, 34, 10, 26, 20])
for c in CAMPS:
    ws.append([c['code'], c.get('name', ''), c.get('parent', ''),
               c.get('start', ''), c.get('end', ''), dur(c),
               ' · '.join(c.get('stations') or []), len(c.get('stations') or []),
               ' · '.join(c.get('tools') or []), c.get('goal', ''), c.get('kpi', ''),
               c.get('target', ''), c.get('budget') or 0, c.get('terms', ''),
               c.get('offer', ''), len(c.get('partners') or []), c.get('notes', ''),
               (c.get('updatedAt') or '')[:10]])
dress(ws)
for r in range(2, ws.max_row + 1):
    ws.cell(r, 1).font = ORNG
    ws.cell(r, 2).font = BOLD
    ws.cell(r, 13).number_format = '#,##0'
NC = ws.max_row

# ------------------------------------------------------------ محطات الحملات
ws = sheet('محطات الحملات', ['كود الحملة', 'اسم الحملة', 'كود المحطة', 'اسم المحطة',
                             'المدينة', 'الشريك في هذه المحطة'],
           [15, 32, 13, 30, 14, 30])
for c in CAMPS:
    pmap = {}
    for p in (c.get('partners') or []):
        if p.get('code'):
            pmap.setdefault(p['code'], []).append(p['name'])
    for code in (c.get('stations') or []):
        ws.append([c['code'], c.get('name', ''), code, stname(code), stcity(code),
                   ' · '.join(pmap.get(code, []))])
dress(ws)
for r in range(2, ws.max_row + 1):
    ws.cell(r, 1).font = ORNG
    ws.cell(r, 3).font = BOLD

# ---------------------------------------------------------------- الشركاء
ws = sheet('الشركاء', ['كود الحملة', 'اسم الحملة', 'الشريك', 'النوع', 'كود المحطة',
                       'اسم المحطة', 'فكرة العرض', 'السعر الاعتيادي', 'سعر الحملة', 'الحالة'],
           [15, 30, 28, 18, 13, 26, 34, 15, 14, 16])
for c in CAMPS:
    for p in (c.get('partners') or []):
        ws.append([c['code'], c.get('name', ''), p.get('name', ''), p.get('type', ''),
                   p.get('code', ''), stname(p.get('code', '')), p.get('offer', ''),
                   p.get('old', ''), p.get('newp', ''), p.get('status', '')])
dress(ws)
for r in range(2, ws.max_row + 1):
    ws.cell(r, 1).font = ORNG
    ws.cell(r, 3).font = BOLD
PR = max(ws.max_row, 60)
dv = DataValidation(type='list', formula1='"%s"' % ','.join(STATUSES), allow_blank=True)
dv.add('J2:J%d' % PR)
ws.add_data_validation(dv)
dvt = DataValidation(type='list', formula1='"%s"' % ','.join(PTYPES), allow_blank=True)
dvt.add('D2:D%d' % PR)
ws.add_data_validation(dvt)

# ----------------------------------------------------------------- الأدوات
ws = sheet('أدوات الحملات', ['كود الحملة', 'اسم الحملة'] + TOOLS,
           [15, 30] + [17] * len(TOOLS))
for c in CAMPS:
    used = set(c.get('tools') or [])
    ws.append([c['code'], c.get('name', '')] + ['✓' if t in used else '' for t in TOOLS])
dress(ws)
for r in range(2, ws.max_row + 1):
    ws.cell(r, 1).font = ORNG
    for col in range(3, 3 + len(TOOLS)):
        ws.cell(r, col).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(r, col).font = Font(name='Arial', size=12, bold=True, color='1B7A4B')

# ---------------------------------------------------------------- الملخّص
MONTHS = ['يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو',
          'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر']
ws = sheet('ملخّص الحملات', ['الشهر', 'كود الحملة', 'اسم الحملة', 'تاريخ البدء',
                            'تاريخ الانتهاء', 'المدة', 'المحطات', 'الشركاء',
                            'الميزانية (ر.س)', 'الهدف'],
           [16, 15, 32, 12, 12, 12, 10, 10, 15, 40])
row = 2
for i, c in enumerate(CAMPS):
    src = i + 2                                    # matching row on الحملات
    st = c.get('start') or ''
    label = ''
    if len(st) >= 7:
        label = '%s %s' % (MONTHS[int(st[5:7]) - 1], st[:4])
    ws.append([label,
               "='الحملات'!A%d" % src, "='الحملات'!B%d" % src, "='الحملات'!D%d" % src,
               "='الحملات'!E%d" % src, "='الحملات'!F%d" % src, "='الحملات'!H%d" % src,
               "='الحملات'!P%d" % src, "='الحملات'!M%d" % src, "='الحملات'!J%d" % src])
    row += 1
ws.append([])
ws.append(['الإجمالي', '', '', '', '', '',
           '=SUM(G2:G%d)' % (row - 1), '=SUM(H2:H%d)' % (row - 1),
           '=SUM(I2:I%d)' % (row - 1), ''])
dress(ws)
for r in range(2, ws.max_row + 1):
    ws.cell(r, 1).font = Font(name='Arial', size=11, bold=True, color=ORANGE)
    ws.cell(r, 9).number_format = '#,##0'
for c in ws[ws.max_row]:
    c.font = BOLD
    c.fill = PatternFill('solid', fgColor=SAND)

# ------------------------------------------------------------ خطة واش واي
WASH = [c for c in CAMPS
        if any('واش واي' in (p.get('name') or '') for p in (c.get('partners') or []))
        or 'واش واي' in (c.get('name') or '') or 'واش واي' in (c.get('goal') or '')]
ws = sheet('خطة واش واي', ['كود الحملة', 'اسم الحملة', 'تاريخ البدء', 'المحطات',
                          'ميزانية الحملة (ر.س)', 'المصروف حتى هنا', 'الباقي من الميزانية'],
           [15, 34, 12, 26, 20, 18, 22])
ws.column_dimensions['I'].width = 20
ws.column_dimensions['J'].width = 16
first = 2
for c in WASH:
    ws.append([c['code'], c.get('name', ''), c.get('start', ''),
               ' · '.join(c.get('stations') or []), c.get('budget') or 0, None, None])
last = ws.max_row
for r in range(first, last + 1):
    ws.cell(r, 6).value = '=SUM($E$%d:E%d)' % (first, r)
    ws.cell(r, 7).value = '=$J$1-F%d' % r
    for col in (5, 6, 7):
        ws.cell(r, col).number_format = '#,##0'
dress(ws, 2, max_col=7)
for r in range(first, last + 1):
    ws.cell(r, 1).font = ORNG
ws['I1'] = 'إجمالي الميزانية'
ws['J1'] = 150000
ws['I2'] = 'المدة (أشهر)'
ws['J2'] = 5
ws['I3'] = 'الشريك'
ws['J3'] = 'واش واي'
ws['I4'] = 'المصروف'
ws['I5'] = 'الباقي'
ws['I6'] = 'حدّ الشهر'
ws['J4'] = '=SUM(E%d:E%d)' % (first, last)
ws['J5'] = '=J1-J4'
ws['J6'] = '=IF(J2>0,J1/J2,0)'
for a in ('J1', 'J4', 'J5', 'J6'):
    ws[a].number_format = '#,##0'
    ws[a].font = BOLD
for a in ('I1', 'I2', 'I3', 'I4', 'I5', 'I6'):
    ws[a].font = BOLD
    ws[a].fill = PatternFill('solid', fgColor=SAND)
    ws[a].border = BOX
for a in ('J1', 'J2', 'J3', 'J4', 'J5', 'J6'):
    ws[a].border = BOX
ws.conditional_formatting.add(
    'G%d:G%d' % (first, last),
    CellIsRule(operator='lessThan', formula=['0'],
               fill=PatternFill('solid', fgColor='F5C6BA'), font=Font(bold=True, color='A33A22')))
ws.conditional_formatting.add(
    'J5', CellIsRule(operator='lessThan', formula=['0'],
                     fill=PatternFill('solid', fgColor='F5C6BA'),
                     font=Font(bold=True, color='A33A22')))

# ------------------------------------------------------------- محطات درب
ws = sheet('محطات درب', ['كود المحطة', 'اسم المحطة', 'المدينة', 'الحالة', 'التشغيل',
                        'المستأجرون', 'وحدات', 'مؤجّرة', 'شاغرة'],
           [13, 34, 16, 14, 16, 60, 9, 9, 9])
for s in D['stations']:
    ws.append([s['code'], s['name'], s.get('city', ''), s.get('status', ''), s.get('op', ''),
               ' · '.join(s.get('tenants') or []), s.get('units', ''),
               s.get('rented', ''), s.get('vacant', '')])
dress(ws)
for r in range(2, ws.max_row + 1):
    ws.cell(r, 1).font = ORNG
ws.auto_filter.ref = 'A1:I%d' % ws.max_row

# --------------------------------------------------------------- القوائم
ws = sheet('القوائم', ['الأدوات', 'أنواع الشركاء', 'حالات الشريك'], [26, 24, 18])
for i in range(max(len(TOOLS), len(PTYPES), len(STATUSES))):
    ws.append([TOOLS[i] if i < len(TOOLS) else '',
               PTYPES[i] if i < len(PTYPES) else '',
               STATUSES[i] if i < len(STATUSES) else ''])
dress(ws)

# ---------------------------------------------------------------- اقرأني
ws = wb['Sheet']
ws.title = 'اقرأني'
ws.sheet_view.rightToLeft = True
ws.column_dimensions['A'].width = 26
ws.column_dimensions['B'].width = 96
rows = [
    ('سجل حملات درب التسويقية', ''),
    ('', ''),
    ('الورقة', 'ماذا فيها'),
    ('الحملات', 'كل حملة في صف واحد بكل حقولها. هذه هي الورقة الأصل — عدّلي فيها.'),
    ('محطات الحملات', 'صف لكل محطة داخل كل حملة، ومعها اسم المحطة ومدينتها وشريكها فيها.'),
    ('الشركاء', 'الشركاء المشاركون بالعرض والسعرين والحالة. النوع والحالة قائمتان منسدلتان.'),
    ('أدوات الحملات', 'جدول علامات: أي أداة استُخدمت في أي حملة.'),
    ('ملخّص الحملات', 'مرتّب بحسب شهر الانطلاق، ويقرأ من ورقة «الحملات» — ما تغيّرينه هناك يظهر هنا.'),
    ('خطة واش واي', 'اكتبي ميزانية كل حملة في العمود «ميزانية الحملة»، فيُحسب المصروف والباقي تلقائيًا، ويحمرّ الرقم لو تجاوز الميزانية.'),
    ('محطات درب', 'محطات درب الـ١٢٧ بمستأجريها — مرشَّحة، للبحث عن كود أي محطة.'),
    ('القوائم', 'الأدوات وأنواع الشركاء وحالاتهم — مرجع القوائم المنسدلة.'),
    ('', ''),
    ('كيف تفتحينها في درايف', 'ارفعي الملف إلى Google Drive، ثم بزر الفأرة الأيمن ← فتح باستخدام ← Google Sheets، ثم ملف ← حفظ كـ Google Sheets.'),
    ('الصيغ', 'تُنقل كما هي وتبقى حيّة بعد التحويل: الملخّص يقرأ من «الحملات»، وخطة واش واي تطرح من الإجمالي.'),
]
for a, b in rows:
    ws.append([a, b])
ws['A1'].font = Font(name='Arial', size=16, bold=True, color=DEEP)
for c in ws[3]:
    c.font = HEAD
    c.fill = PatternFill('solid', fgColor=DEEP)
for r in range(4, ws.max_row + 1):
    ws.cell(r, 1).font = BOLD
    ws.cell(r, 1).alignment = MID
    ws.cell(r, 2).alignment = WRAP

wb.move_sheet('اقرأني', offset=-wb.sheetnames.index('اقرأني'))
OUT = '/home/user/-/darb-campaigns.xlsx'
wb.save(OUT)
print('sheets:', wb.sheetnames)
print('campaigns:', NC - 1, '| wash rows:', len(WASH), '| stations:', len(D['stations']))
print('saved:', OUT)
