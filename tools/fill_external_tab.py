# -*- coding: utf-8 -*-
"""يملأ تبويب «الشركاء الخارجيون» ببيانات المنشآت، ويضيف تراجع/إعادة وحذف الصفوف."""
import openpyxl, re, html, collections, json, os

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(BASE, "darb-external-partners.xlsx")
LOCS = os.path.join(BASE, "darb-station-locations.xlsx")
SRC = OUT = os.path.join(BASE, "darb-five-stations-analysis.html")
KEEP = ["MK007", "MK017", "MK002", "MK023", "MK019"]

#  أسماء المحطات في ملف المنشآت → الأكواد
NAME2CODE = {"العمرة الجديدة": "MK007", "عرفات الشوقية": "MK017", "المعيصم": "MK002",
             "بن درويش": "MK023", "عرفات الشرائع": "MK019"}
#  أيقونة وسبب أهمية كل فئة كما وردت في الملف
CATMETA = {
    "مكاتب تأجير سيارات": ("🚗", "أساطيل تأجير تعبّئ يوميًا — عقد شهري بفاتورة موحّدة وخصم كميات"),
    "خدمات سيارات":       ("🔧", "المحطة هي التزود الطبيعي لعملائهم ولسيارات المركز نفسها"),
    "شركات":              ("🏢", "سيارات شركة وموظفون يتنقلون يوميًا — عقد تعبئة وبطاقة سائق"),
    "مدارس":              ("🎓", "حافلات نقل الطلاب تعبّئ بانتظام في مواعيد ثابتة"),
}
COLS = ["المنشأة", "الهاتف", "الحالة", "الخدمة المقترحة", "ملاحظة"]
STATUSES = ["محتمل", "تم التواصل", "عميل حالي", "مستبعد"]

E = lambda t: html.escape(str(t), quote=True)

# ═══════════ 1. البيانات ═══════════
ws = openpyxl.load_workbook(XLSX, data_only=True)["Sheet1"]
RECS = collections.defaultdict(lambda: collections.defaultdict(list))
CATS_SEEN = []
for r in list(ws.iter_rows(values_only=True))[1:]:
    if not r or not r[0]:
        continue
    st, cat = str(r[0]).strip(), str(r[1]).strip()
    code = NAME2CODE.get(st)
    assert code, f"محطة غير معروفة: {st}"
    phone = str(r[3]).strip() if r[3] else ""
    if phone in ("غير متوفر", "None", ""):
        phone = ""
    RECS[code][cat].append({"name": str(r[2]).strip(), "phone": phone})
    if cat not in CATS_SEEN:
        CATS_SEEN.append(cat)

ws2 = openpyxl.load_workbook(LOCS, data_only=True)["الورقة1"]
LOC = {}
for r in list(ws2.iter_rows(values_only=True))[1:]:
    if not r or not r[0]:
        continue
    m = re.search(r"(-?\d+\.\d+)\s*,\s*(-?\d+\.\d+)", str(r[1]))
    u = re.search(r"(https?://\S+)", str(r[1]))
    LOC[str(r[0]).strip()] = dict(lat=float(m.group(1)), lng=float(m.group(2)),
                                  url=u.group(1) if u else "")

# ═══════════ 2. تقسيم الصفحات ═══════════
src = open(SRC, encoding="utf-8").read()
pages_head, pages_all = src.split('<main class="wrap" id="pages">', 1)
blocks = re.split(r'(?=<div class="pgview")', pages_all)
lead, blocks = blocks[0], blocks[1:]
PG, ORDER = {}, []
for b in blocks:
    k = re.search(r'id="pg-([\w-]+)"', b).group(1)
    PG[k] = b
    ORDER.append(k)

CS_COUNT = {c: len(re.findall(r'<tr><td>\d{4}-\d\d-\d\d</td>', PG[c + "-cs"])) for c in KEEP}
PT_COUNT = {c: int(re.search(r'الشركاء داخل المحطة</div><div class="kv">(\d+)',
                             PG[c + "-partners"]).group(1)) for c in KEEP}
XT_COUNT = {c: sum(len(v) for v in RECS[c].values()) for c in KEEP}


def tabs_of(code, active):
    t = [("", "التحليل الكامل"), ("/monthly", "المبيعات الشهرية"), ("/daily", "المبيعات اليومية"),
         ("/targets", "المستهدفات"), ("/cs", "استفسارات العملاء"),
         ("/partners", "الشركاء عبر اليوم"), ("/external", "الشركاء الخارجيون")]
    return '<div class="tabs">' + "".join(
        f'<a class="tab{" on" if suf == active else ""}" href="#/{code}{suf}">{lbl}'
        + (f' <b class="tcount">{CS_COUNT[code]}</b>' if suf == "/cs" else "")
        + (f' <b class="tcount">{PT_COUNT[code]}</b>' if suf == "/partners" else "")
        + (f' <b class="tcount">{XT_COUNT[code]}</b>' if suf == "/external" else "")
        + "</a>" for suf, lbl in t) + "</div>"


def status_cell(v="محتمل"):
    return f'<span class="cls c-un xstate" title="اضغط لتغيير الحالة">{E(v)}</span>'


def table_of(items):
    head = "".join(f"<th>{c}</th>" for c in COLS)
    body = "".join(
        f'<tr><td><b>{E(x["name"])}</b></td>'
        f'<td>{E(x["phone"]) if x["phone"] else "—"}</td>'
        f"<td>{status_cell()}</td><td>—</td><td>—</td></tr>" for x in items)
    return ('<div class="ntable xtbl"><div class="tscroll"><table>'
            f"<thead><tr>{head}</tr></thead><tbody>{body}</tbody></table></div></div>")


# ═══════════ 3. إعادة بناء صفحة الشركاء الخارجيين ═══════════
for c in KEEP:
    L, byc = LOC[c], RECS[c]
    total = XT_COUNT[c]
    with_phone = sum(1 for v in byc.values() for x in v if x["phone"])
    cats_here = [k for k in CATS_SEEN if byc.get(k)]
    top = max(cats_here, key=lambda k: len(byc[k])) if cats_here else "—"

    kpis = (
        '<div class="skpis" style="grid-template-columns:repeat(6,1fr)">'
        f'<div class="kpi hot"><div class="kl">منشآت مرصودة</div><div class="kv">{total}</div>'
        f'<div class="kn">حول المحطة — قابلة للإضافة والحذف</div></div>'
        f'<div class="kpi"><div class="kl">لديها رقم تواصل</div><div class="kv">{with_phone}'
        f'<small> من {total}</small></div>'
        f'<div class="kn">{with_phone/total*100:.0f}٪ جاهزة للتواصل المباشر</div></div>'
        f'<div class="kpi"><div class="kl">الفئات</div><div class="kv">{len(cats_here)}</div>'
        f'<div class="kn">' + " · ".join(f"{k} {len(byc[k])}" for k in cats_here) + "</div></div>"
        f'<div class="kpi"><div class="kl">الفئة الأكبر</div>'
        f'<div class="kv" style="font-size:15px">{E(top)}</div>'
        f'<div class="kn">{len(byc[top]) if top in byc else 0} منشأة</div></div>'
        f'<div class="kpi"><div class="kl">إحداثيات المحطة</div>'
        f'<div class="kv" style="font-size:14px">{L["lat"]:.5f}<br>{L["lng"]:.5f}</div>'
        f'<div class="kn"><a href="{E(L["url"])}" target="_blank" rel="noopener">افتح في الخرائط ↗</a></div></div>'
        f'<div class="kpi"><div class="kl">حالة التواصل</div><div class="kv">—</div>'
        f'<div class="kn">حدّثها بالضغط على وسم الحالة في كل صف</div></div>'
        "</div>")

    sections = ""
    for cat in cats_here:
        ico, why = CATMETA.get(cat, ("📌", ""))
        sections += (
            f'<div class="sec-h" style="margin-top:20px"><h2>{ico} {E(cat)}</h2>'
            f'<span>{len(byc[cat])} منشأة · {why}</span></div>'
            + table_of(byc[cat]))

    mini = re.search(r'<div class="mini-head">.*?</div>\s*(?=<div class="skpis")',
                     PG[c + "-monthly"], re.S).group(0)
    nav = re.search(r'<div class="pgnav">.*?</select>\s*</div>', PG[c], re.S).group(0)
    ez = re.search(r'id="pg-' + c + r'-external"', PG[c + "-external"])
    ezn = re.search(r'data-ez="(z\d+)"[^>]*id="pg-' + c + '-external"', PG[c + "-external"]).group(1)
    title = re.search(r'data-title="([^"]*)"', PG[c]).group(1).split(" · ")[0]

    PG[c + "-external"] = (
        f'<div class="pgview" data-ez="{ezn}" id="pg-{c}-external" '
        f'data-title="{title} · الشركاء الخارجيون" hidden>'
        + nav + tabs_of(c, "/external") + mini + kpis
        + '<div class="dnote" style="background:#F0F5FA;border:1px solid #CBDDEB">🎯 '
          '<b>الشريك الخارجي:</b> منشأة <b>خارج</b> المحطة تستطيع المحطة أن تخدمها — '
          'عقد تعبئة أسطول، خصم كميات، بطاقة سائق، أو خدمات مساندة. تختلف عن '
          '<b>الشركاء داخل المحطة</b> في التبويب المجاور، وعن <b>المنافسين</b> في التحليل الكامل. '
          '<b>الأعمدة قابلة للتحرير بالكامل</b>: فعّل وضع التحرير، واكتب في أي خلية، '
          'واحذف أي صف بزر × عند اسمه، وأضف صفًا بزر «＋ صف»، وتراجع بزر «↶» في الشريط.</div>'
        + sections
        + '<div class="dnote">📐 <b>المصدر:</b> ملف المنشآت المرفق — الأسماء وأرقام التواصل كما '
          'وردت فيه دون تعديل. أعمدة <b>الحالة</b> و<b>الخدمة المقترحة</b> و<b>ملاحظة</b> فارغة '
          'ومتروكة لكم. الملف لا يتضمن إحداثيات المنشآت فلا يوجد عمود مسافة؛ عند توفّرها '
          'أضيف المسافة المحسوبة من إحداثيات المحطة.</div>'
        + '<div class="pgnav" style="margin-top:4px"><div class="nvl">'
          f'<a class="hb" href="#/">⌂ جميع المحطات</a>'
          f'<a href="#/{c}/partners">← الشركاء عبر اليوم</a></div></div></div>')

#  تحديث شريط التبويبات في بقية الصفحات (تغيّر العدّاد)
for c in KEEP:
    for suf, key in (("", c), ("/monthly", c + "-monthly"), ("/daily", c + "-daily"),
                     ("/targets", c + "-targets"), ("/cs", c + "-cs"),
                     ("/partners", c + "-partners")):
        p, k = re.subn(r'<div class="tabs">.*?</div>', tabs_of(c, suf), PG[key], count=1, flags=re.S)
        assert k == 1, key
        PG[key] = p

pages_all = lead + "".join(PG[k] for k in ORDER)

# ═══════════ 4. تراجع/إعادة + حذف الصفوف ═══════════
CSS = """
<style id="undo-css">
/* ── تراجع/إعادة وحذف الصفوف ── */
#edbar .edhist{border-color:var(--line2);min-width:34px;font-size:15px;line-height:1}
#edbar .edhist:disabled{opacity:.35;cursor:default;border-color:var(--line)}
#edbar .edhist:disabled:hover{border-color:var(--line);color:var(--ink2)}
html.editing table tbody tr td:first-child{position:relative}
.bdrowx{position:absolute;inset-inline-start:-2px;inset-block-start:50%;transform:translateY(-50%);
  width:17px;height:17px;border-radius:50%;background:#FBF0ED;border:1px solid #E3BEB4;color:#A6432E;
  font-size:12px;font-weight:800;line-height:15px;text-align:center;cursor:pointer;opacity:0;
  transition:.12s;user-select:none;-webkit-user-select:none}
html.editing tr:hover .bdrowx{opacity:1}
.bdrowx:hover{background:#C0503A;color:#fff;border-color:#C0503A}
html.editing table tbody tr td:first-child{padding-inline-start:22px}
.xtbl td{white-space:normal}
.xtbl td:first-child{min-width:230px}
</style>
"""

JS = r"""
<script id="undo-js">
/* ═══ تراجع وإعادة على مستوى التعديل + حذف أي صف ═══
   لقطة لكل منطقة تحرير بعد كل تغيير مستقر، ومكدّسان للتراجع والإعادة.
   لا يلمس الملف نفسه — يتراجع عن التعديل داخل الصفحة فقط.            */
(function(){
  var root=document.documentElement;
  var ZONES=[].slice.call(document.querySelectorAll('[data-ez]'));
  var undoS=[], redoS=[], LAST={}, timer=null, applying=false;

  function snap(z){
    var c=z.cloneNode(true);
    c.querySelectorAll('[data-builder]').forEach(function(n){n.remove();});
    c.querySelectorAll('[data-bddrag]').forEach(function(n){n.removeAttribute('data-bddrag');});
    c.querySelectorAll('[contenteditable]').forEach(function(n){n.removeAttribute('contenteditable');});
    c.querySelectorAll('[spellcheck]').forEach(function(n){n.removeAttribute('spellcheck');});
    return c.innerHTML;
  }
  ZONES.forEach(function(z){LAST[z.dataset.ez]=snap(z);});

  function commit(){
    if(applying)return;
    var ch=[];
    ZONES.forEach(function(z){
      var k=z.dataset.ez, h=snap(z);
      if(h!==LAST[k]){ ch.push({ez:k,before:LAST[k],after:h}); LAST[k]=h; }
    });
    if(ch.length){ undoS.push(ch); redoS.length=0; if(undoS.length>60)undoS.shift(); paint(); }
  }
  function apply(entry,side){
    applying=true;
    entry.forEach(function(e){
      var z=document.querySelector('[data-ez="'+e.ez+'"]');
      if(!z)return;
      z.innerHTML=e[side];
      LAST[e.ez]=e[side];
      if(root.classList.contains('editing')){
        z.setAttribute('contenteditable','true'); z.setAttribute('spellcheck','false');
      }
      z.dispatchEvent(new Event('input',{bubbles:true}));   /* ليحفظ المحرّر */
    });
    setTimeout(function(){applying=false;paint();},0);
  }
  function undo(){ var e=undoS.pop(); if(!e)return; apply(e,'before'); redoS.push(e); }
  function redo(){ var e=redoS.pop(); if(!e)return; apply(e,'after');  undoS.push(e); }

  ZONES.forEach(function(z){
    z.addEventListener('input',function(){ clearTimeout(timer); timer=setTimeout(commit,700); });
  });

  /* الأزرار داخل شريط التحرير */
  var bar=document.getElementById('edbar');
  var bu=document.createElement('button'), br=document.createElement('button');
  bu.className='edhist'; bu.id='edUndo'; bu.textContent='↶'; bu.title='تراجع (Ctrl+Z)';
  br.className='edhist'; br.id='edRedo'; br.textContent='↷'; br.title='إعادة (Ctrl+Shift+Z)';
  var anchor=bar.querySelector('#edStat');
  bar.insertBefore(bu,anchor); bar.insertBefore(br,anchor);
  bu.addEventListener('click',function(){clearTimeout(timer);commit();undo();});
  br.addEventListener('click',redo);
  function paint(){
    bu.disabled=!undoS.length; br.disabled=!redoS.length;
    bu.title='تراجع (Ctrl+Z)'+(undoS.length?' — '+undoS.length+' خطوة':'');
    br.title='إعادة (Ctrl+Shift+Z)'+(redoS.length?' — '+redoS.length+' خطوة':'');
  }
  paint();

  document.addEventListener('keydown',function(e){
    if(!(e.ctrlKey||e.metaKey))return;
    var k=e.key.toLowerCase();
    if(k==='z'&&!e.shiftKey){e.preventDefault();clearTimeout(timer);commit();undo();}
    else if((k==='z'&&e.shiftKey)||k==='y'){e.preventDefault();redo();}
  });

  /* حذف أي صف — زر × يظهر عند المرور على الصف في وضع التحرير */
  function rowChrome(){
    document.querySelectorAll('.bdrowx').forEach(function(n){n.remove();});
    if(!root.classList.contains('editing'))return;
    ZONES.forEach(function(z){
      z.querySelectorAll('table tbody tr').forEach(function(tr){
        var td=tr.cells[0]; if(!td||td.querySelector('.bdrowx'))return;
        var x=document.createElement('span');
        x.className='bdrowx'; x.setAttribute('data-builder',''); x.contentEditable='false';
        x.title='حذف هذا الصف'; x.textContent='×';
        td.appendChild(x);
      });
    });
  }
  document.addEventListener('click',function(e){
    var x=e.target.closest&&e.target.closest('.bdrowx'); if(!x)return;
    e.preventDefault(); e.stopPropagation();
    var tr=x.closest('tr'), z=tr.closest('[data-ez]');
    tr.remove();
    if(z)z.dispatchEvent(new Event('input',{bubbles:true}));
    clearTimeout(timer); commit();
  },true);
  /* الضغط على خلية فارغة (—) يحدّد محتواها فتحلّ الكتابة محلّه مباشرة */
  document.addEventListener('click',function(e){
    if(!root.classList.contains('editing'))return;
    var td=e.target.closest&&e.target.closest('td');
    if(!td||!td.closest('[data-ez]')||td.textContent.trim()!=='—')return;
    var r=document.createRange(); r.selectNodeContents(td);
    var sel=window.getSelection(); sel.removeAllRanges(); sel.addRange(r);
  });

  new MutationObserver(rowChrome).observe(root,{attributes:true,attributeFilter:['class']});
  rowChrome();
})();
</script>
"""

doc = pages_head + '<main class="wrap" id="pages">' + pages_all
doc = doc.replace("</head>", CSS + "</head>", 1)
doc = doc.replace("</body>", JS + "</body>", 1)
doc = re.sub(r'data-docid="[^"]*"', 'data-docid="darb-5st-xfill-v1"', doc, count=1)
open(OUT, "w", encoding="utf-8").write(doc)

print("تم · الحجم:", round(len(doc.encode()) / 1024), "KB · الفئات:", CATS_SEEN)
for c in KEEP:
    print(f"  {c}: {XT_COUNT[c]} منشأة — "
          + " · ".join(f"{k} {len(v)}" for k, v in RECS[c].items()))
